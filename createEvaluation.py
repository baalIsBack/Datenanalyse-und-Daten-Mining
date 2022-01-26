from openpyxl import *

#Globale Variable und Einstellungsmoeglichkeit fuer die Anzahl der Clients
#ACHTUNG: EXPONENTIALFALLE
numberOfClients = 10
numberOfRounds = 20 #Die Anzahl der Runden pro Run
numberOfRuns = 15 #Die Anzahl der Runs (Wie oft wird das Programm ausgeführt)
path = 'Evaluation/FLStandard/'

def main(): #Erstellung von Workbook Versionen
    if input("Pls confirm the override / creation of the Evaluations Sheets (y) : ").lower() == "y":
        addVersionWorkbook('Version1')
        addVersionWorkbook('Version2')
        addVersionWorkbook('Version3')
        print("Sheets are created / replaced")
    else:
        print("Program aborted")

def addVersionWorkbook(name): #Erstellung von einem Workbook
    if (numberOfClients > 20):
        raise AssertionError

    wb = Workbook()

    addEvaluationSheet("Average loss during Training", wb)
    addEvaluationSheet('Global Test loss', wb)
    addEvaluationSheet('Global Accuracy', wb)
    for clientid1 in range(numberOfClients):
        for clientid2 in range(numberOfClients):
            addEvaluationSheet('CM ' + str(clientid2) + str(clientid1), wb)

    summary(name, wb)

    wb.active = wb['Sheet']
    wb.save(path+name+'.xlsx')

def addEvaluationSheet(name, wb): #Erstellung von Worksheets
    ws = wb.create_sheet(name)
    wb.active = ws
    ws['A1'].value = "Average"
    #Erstellung der Überschriften
    pointer = 1
    for row in ws.iter_rows(1, 1, 2, numberOfRuns+1):
        for cell in row:
            cell.value = "Run" + str(pointer)
            pointer += 1
    # Erstellung der Durchschnittswerte
    pointer = 2
    for row in ws.iter_rows(2, numberOfRounds+1, 1, 1):
        for cell in row:
            cell.value = "=Average(B" + str(pointer) + ":P" + str(pointer) + ")"
            pointer += 1

def summary(version, wb):
    #Auswahl der richtigen Stelle
    ws = wb['Sheet']

    #Erstellung der Spaltenüberschriften
    ws['A1'].value = 'Average loss during Training'
    ws['B1'].value = 'Global Test loss'
    ws['C1'].value = 'Global Accuracy'
    ws['E1'].value = 'Accuracy Matritzen'

    #Übernahme drei Main Werte Werte
    for index in range(2,2+numberOfRounds):
        ws['A'+str(index)] = '=\'Average loss during Training\'!A'+str(index)
        ws['B' + str(index)] = '=\'Global Test loss\'!B' + str(index)
        ws['C' + str(index)] = '=\'Global Accuracy\'!C' + str(index)

    #Anpassung der Zellenbreite, einfach nicht hinterfragen, dass das ganze mit 7 umgerechnet werden muss
    ws.column_dimensions['A'].width = 200/7
    ws.column_dimensions['B'].width = 120/7
    ws.column_dimensions['C'].width = 120/7

    #Übernahme der Accuracy Matritzen
    #Für jede Matrix aller Runs wird jede Spalte Aufgebaut
    pointerx = 0
    for pointerRounds in range(1,numberOfRounds+1):#Matrix
        ws['E'+str(3+(pointerRounds-1)*(numberOfClients+2))].value = 'Round '+str(pointerRounds)
        for pointery in range(10): #Spalte
            for row in ws.iter_rows(4+(pointerRounds-1)*(numberOfClients+2), (numberOfClients+3)+(pointerRounds-1)*(numberOfClients+2), pointery+5, pointery+5):#Reihe in der 1D Spalte
                for cell in row:
                    cell.value = '=\'CM '+str(pointerx)+str(pointery)+'\'!A'+str(2+pointerRounds-1)
                    pointerx +=1
            pointerx = 0


if __name__ == '__main__':
    main()
