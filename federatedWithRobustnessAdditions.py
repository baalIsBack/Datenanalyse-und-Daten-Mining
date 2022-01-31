import _pickle
import os
import time

import numpy as np
import numpy.random
import torch
import torch.nn as nn
import torch.nn.functional
import torch.optim
from sklearn.model_selection import train_test_split
from torch.utils.data import DataLoader
from torchvision import transforms
from torch.optim.lr_scheduler import StepLR
from torchvision.datasets import MNIST
from torchvision.utils import save_image
from tqdm import tqdm
from torch.utils.data import Dataset
import skimage
from openpyxl import *
from openpyxl.utils import get_column_letter

# definition of Parameters
numberOfEpochs = 3
batchSize = 12
learningrate = 0.001
numberOfClients = 10
numberOfSelectedClients = 4
numberOfRounds =20
training=False
autoencoder_disabled=False


# documentation
path = "Evaluation/FLStandard/"



def getOther(n):
    x = np.random.randint(10)
    if x == n:
        return getOther(n)
    return x

class Dataset_RandomOther(Dataset):
    def __init__(self, transform):
        self.mnist = MNIST(root='./rootMNIST', train=True, download=True, transform=transform)

    def __len__(self):
        return len(self.mnist)

    def __getitem__(self, idx):
        it = list(self.mnist[idx])
        it[-1] = getOther(it[-1])
        return it

class Dataset_Shift(Dataset):
    def __init__(self, transform):
        self.mnist = MNIST(root='./rootMNIST', train=True, download=True, transform=transform)

    def __len__(self):
        return len(self.mnist)

    def __getitem__(self, idx):
        it = list(self.mnist[idx])
        it[-1] = it[-1]+1 % 10
        return it

# definition of the Neural Network Model with 4 Layers
def NeuralNetwork():
    layers = []
    layers.append(nn.Conv2d(1, 16, 3, 1))
    layers.append(nn.ReLU())
    layers.append(nn.Flatten(1))
    layers.append(nn.Linear(26 * 26 * 16, 128))
    layers.append(nn.ReLU())
    layers.append(nn.Dropout(0.25))
    layers.append(nn.Linear(128, 64))
    layers.append(nn.ReLU())
    layers.append(nn.Dropout(0.5))
    layers.append(nn.Linear(64, 10))
    return nn.Sequential(*layers)


"------------------Test Functions, starting with test Function for original Images, then for Images with Noise ( with and without Image pre-processing Autoencoder)------------------------------------"

# testing accuracy on global model after each round
def test(model, device, test_loader):
    correctly_classified = 0
    loss = 0
    with torch.no_grad():
        for image, target in test_loader:
            img, target = image.to(device), target.to(device)
            outputs = model(img)
            _, predicted = torch.max(outputs, 1)
            loss += nn.CrossEntropyLoss()(outputs, predicted).item()
            correctly_classified += (predicted == target).sum().item()
    loss /= len(test_loader.dataset)
    accuracy = 100.0 * correctly_classified / len(test_loader.dataset)

    return loss, accuracy


# testing accuracy on global model after each round with images perturbed with SaltPepper Noise Without Image Preprocessing by Autoencoder
def test_SaltPepper(model, device, test_loader):
    correctly_classified = 0
    loss = 0
    with torch.no_grad():
        for image, target in test_loader:
            target =  target.to(device)
            images_with_added_noise =torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True)).to(device)
            outputs = model(images_with_added_noise)
            _, predicted = torch.max(outputs, 1)
            loss += nn.CrossEntropyLoss()(outputs, predicted).item()
            correctly_classified += (predicted == target).sum().item()

    loss /= len(test_loader.dataset)
    accuracy = 100.0 * correctly_classified / len(test_loader.dataset)

    return loss, accuracy

def test_HardSaltPepper(model, device, test_loader):
    correctly_classified = 0
    loss = 0
    with torch.no_grad():
        for image, target in test_loader:
            target =  target.to(device)
            option1 = torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True))
            option2 = torch.tensor(skimage.util.random_noise(
                     torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
            option3 = torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                 , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                 , mode='s&p', salt_vs_pepper=0.5, clip=True))
            option4 = torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                 , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                 , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
            option5 = torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(
                     torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
            option6 = torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(
                      torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
            chosenPerturbation = numpy.random.choice([1, 2, 3, 4, 5, 6])

            if (  chosenPerturbation == 1):
                randomPerturbation = option1
            elif (  chosenPerturbation == 2):
                randomPerturbation = option2
            elif (  chosenPerturbation == 3):
                randomPerturbation = option3
            elif (  chosenPerturbation == 4):
                randomPerturbation = option4
            elif (  chosenPerturbation == 5):
                randomPerturbation = option5
            elif (  chosenPerturbation == 6):
                randomPerturbation = option6

            images_with_added_noise = randomPerturbation.to(device)
            outputs = model(images_with_added_noise)
            _, predicted = torch.max(outputs, 1)
            loss += nn.CrossEntropyLoss()(outputs, predicted).item()
            correctly_classified += (predicted == target).sum().item()

    loss /= len(test_loader.dataset)
    accuracy = 100.0 * correctly_classified / len(test_loader.dataset)

    return loss, accuracy


def test_HardSaltPepperWithAutoencoder(model, device, test_loader,autoencoder):


    correctly_classified = 0
    loss = 0
    with torch.no_grad():
        for image, target in test_loader:
            target =  target.to(device)
            option1 = torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True)).to(
                device)
            option2 = torch.tensor(skimage.util.random_noise(
                torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True))
                , mode='s&p', salt_vs_pepper=0.5, clip=True))
            option3 = torch.tensor(skimage.util.random_noise(
                torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True))
                    , mode='s&p', salt_vs_pepper=0.5, clip=True))
                , mode='s&p', salt_vs_pepper=0.5, clip=True))
            option4=torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
            option5=torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
            option6=torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(
                    torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
            chosenPerturbation = numpy.random.choice([1, 2, 3, 4, 5, 6])

            if (chosenPerturbation == 1):
                randomPerturbation = option1
            elif (chosenPerturbation == 2):
                randomPerturbation = option2
            elif (chosenPerturbation == 3):
                randomPerturbation = option3
            elif (chosenPerturbation == 4):
                randomPerturbation = option4
            elif (chosenPerturbation == 5):
                randomPerturbation = option5
            elif (chosenPerturbation == 6):
                randomPerturbation = option6

            images_with_added_noise = randomPerturbation.to(device)
            image = autoencoder(images_with_added_noise);
            outputs = model(image)
            _, predicted = torch.max(outputs, 1)
            loss += nn.CrossEntropyLoss()(outputs, predicted).item()
            correctly_classified += (predicted == target).sum().item()

    loss /= len(test_loader.dataset)
    accuracy = 100.0 * correctly_classified / len(test_loader.dataset)

    return loss, accuracy



# testing accuracy on global model after each round with images perturbed with SaltPepper Noise With Image Preprocessing by Autoencoder
def test_SaltPepperWithAutoEncoder(model, device, test_loader,autoencoder):
    correctly_classified = 0
    loss = 0
    with torch.no_grad():
        for image, target in test_loader:
            target = target.to(device)
            images_with_added_noise =torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True)).to(device)
            #Remove Comments to save Noisy/ Denoised Images.
            #if not os.path.exists('TestImages'):
            #    os.makedirs('TestImages')
            #saveImageCostum( images_with_added_noise.cpu().data, name='./TestImages/noisy{}.png'.format(i))
            image = autoencoder( images_with_added_noise)
            #saveImageCostum(image.cpu().data, name='./TestImages/denoised{}.png'.format(i))
            outputs = model(image)
            _, predicted = torch.max(outputs, 1)
            loss += nn.CrossEntropyLoss()(outputs, predicted).item()
            correctly_classified += (predicted == target).sum().item()

    loss /= len(test_loader.dataset)
    accuracy = 100.0 * correctly_classified / len(test_loader.dataset)

    return loss, accuracy



#TODO:Test Methods for Gaussian Noise,Speckle Noise
"---------------------------------------------Train Functions, starting with train Function for original Images, then for Images with Noise---------------------------------------------------------------------"


# trains model of client on client data
def train_client(model, device, optimizer, train_loader, epoch):
    for epochX in range(epoch):
        i=0
        for batch, (image, target) in enumerate(train_loader):
            img, target = image.to(device), target.to(device)
            optimizer.zero_grad()
            output = model(img)
            loss = nn.CrossEntropyLoss()(output, target)
            loss.backward()
            optimizer.step()
            i+=1
    return loss.item()

#TrainClient on Salt Pepper Noise
def train_client_SaltPepper(model, device, optimizer, train_loader, epoch):
    for epochX in range(epoch):
        i=0
        for batch, (image, target) in enumerate(train_loader):
            images_with_added_noise = torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True)).to(device)
            optimizer.zero_grad()
            output = model(images_with_added_noise)
            loss = nn.CrossEntropyLoss()(output, target)
            loss.backward()
            optimizer.step()
            i+=1
    return loss.item()

def train_client_HardSaltPepper(model, device, optimizer, train_loader, epoch):
    for epochX in range(epoch):
        i=0
        for batch, (image, target) in enumerate(train_loader):
            images_with_added_noise = torch.tensor(skimage.util.random_noise(
                                      torch.tensor(skimage.util.random_noise(torch.tensor(skimage.util.random_noise(image
                                                                                                , mode='s&p', salt_vs_pepper=0.5,clip=True))
                                                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True))
                                                                                                , mode='s&p', salt_vs_pepper=0.5, clip=True)).to(device)
            optimizer.zero_grad()
            output = model(images_with_added_noise)
            loss = nn.CrossEntropyLoss()(output, target)
            loss.backward()
            optimizer.step()
            i+=1
    return loss.item()

#TODO: Training Methods for Gaussian Noise,Speckle Noise
"---------------------------------------------------------------------------------------------------------------------------------------------------------------"


# aggregates results of the selected and trained client models + takes the mean of the weights to update global model and all client models
def server_aggregate(global_model, client_models,selected_clients):
    # get state of global model
    update_global_model = global_model.state_dict()
    # average the weights of selected clients and update global model
    for weighti in update_global_model.keys():
        update_global_model[weighti] = torch.stack(
            [client_models[selected_clients[i]].state_dict()[weighti].float() for i in range(numberOfSelectedClients)], 0).mean(0)
        global_model.load_state_dict(update_global_model)
    # update the models of all clients before next training
    for model in client_models:
        model.load_state_dict(global_model.state_dict())


"------------------------------------------------------ AutoEncoder Methods Start------------------------------------------------"
def saveImageCostum(img, name):
    img = img.view(img.size(0), 1, 28, 28)
    save_image(img, name)

def Autoencoder():
    layers = []
    #Layers for Encoding
    layers.append(nn.Conv2d(1, 64, kernel_size=3, padding=1))
    layers.append(nn.ReLU())
    layers.append(nn.MaxPool2d(2, 2))
    layers.append(nn.Conv2d(64, 32, kernel_size=3, padding=1))
    layers.append(nn.ReLU())
    layers.append(nn.MaxPool2d(2, 2))
    layers.append(nn.Conv2d(32, 16, kernel_size=3, padding=1))
    layers.append(nn.ReLU())
    layers.append(nn.MaxPool2d(2, 2))
    layers.append(nn.Conv2d(16, 8, kernel_size=3, padding=1))
    layers.append(nn.ReLU())
    layers.append(nn.MaxPool2d(2, 2))

    #Layers for decoding
    layers.append(nn.ConvTranspose2d(8, 8, kernel_size=3, stride=2))
    layers.append(nn.ReLU())
    layers.append(nn.ConvTranspose2d(8, 16, kernel_size=3, stride=2))
    layers.append(nn.ReLU())
    layers.append(nn.ConvTranspose2d(16, 32, kernel_size=2, stride=2))
    layers.append(nn.ReLU())
    layers.append(nn.ConvTranspose2d(32, 64, kernel_size=2, stride=2))
    layers.append(nn.ReLU())
    layers.append(nn.Conv2d(64, 1, kernel_size=3, padding=1))
    layers.append(nn.Sigmoid())

    return nn.Sequential(*layers)






# train for 8 Epochs after regular Training to adapt Autoencoder to our neural classification network  (prevent autoencoder from filtering out too much, which sometimes leads to mispredictions)
def trainAutoEncoderWithNN(net, trainloader, numberOfEpochs,  optimizer, device,NN):
        train_loss = []
        for epoch in range(numberOfEpochs):
            running_loss = 0.0
            for batch, (image, target) in enumerate(trainloader):
                target=target.to(device)
                images_with_added_noise = torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True)).to(device)
                optimizer.zero_grad()
                outputs = net(images_with_added_noise)
                loss1 = nn.MSELoss()(outputs, image.to(device))
                loss2=nn.CrossEntropyLoss()(NN(outputs), target)
                loss=loss1+loss2
                loss.backward()
                optimizer.step()
            loss = running_loss / len(trainloader)
            train_loss.append(loss)
            print('Epoch {} of {}, Train Loss: {:.3f}'.format(
                epoch + 1, numberOfEpochs, loss))
            saveImageCostum(images_with_added_noise.cpu().data, name='./TestImages/noisy{}.png'.format(epoch))
            saveImageCostum(outputs.cpu().data, name='./TestImages/denoised{}.png'.format(epoch))
        return train_loss

#regular Training of Autoencoder
def trainAutoEncoder(net, trainloader, numberOfEpochs,  optimizer, device):
        train_loss = []
        for epoch in range(numberOfEpochs):
            running_loss = 0.0
            for batch, (image, target) in enumerate(trainloader):
                images_with_added_noise = torch.tensor(skimage.util.random_noise(image, mode='s&p', salt_vs_pepper=0.5, clip=True)).to(device)
                optimizer.zero_grad()
                outputs = net(images_with_added_noise)
                loss = nn.MSELoss()(outputs, image.to(device))
                loss.backward()
                optimizer.step()
                running_loss += loss.item()
            loss = running_loss / len(trainloader)
            train_loss.append(loss)
            print('Epoch {} of {}, Train Loss: {:.3f}'.format(
                epoch + 1, numberOfEpochs, loss))
            saveImageCostum(images_with_added_noise.cpu().data, name='./TestImages/noisy{}.png'.format(epoch))
            saveImageCostum(outputs.cpu().data, name='./TestImages/denoised{}.png'.format(epoch))
        return train_loss



def prepareAutoEncoder(trainloader,device,NN):

    autoencoder = Autoencoder()
    optimizer = torch.optim.Adam(autoencoder.parameters(), lr=0.001)
    autoencoder.to(device)
    if not os.path.exists('TestImages'):
        os.makedirs('TestImages')
    trainAutoEncoder(autoencoder, trainloader, 20,optimizer,device)
    trainAutoEncoderWithNN(autoencoder,trainloader,8,optimizer,device,NN)
    torch.save(autoencoder, ("./AutoEncoder"))

    return autoencoder
"------------------------------------------------------ AutoEncoder End---------------------------------------------------------------"


#Documentation
def getColumn(ws):
    pointer = 2
    for row in ws.iter_rows(2, 2, 2, 100):
        for cell in row:
            if cell.value is None:
                return pointer
            else:
                pointer += 1


def main():
    # Gathering Inputs
    print('Please enter a Number to select a Version:')
    print('Version 1: One worker with mainly 0es')
    print('Version 2: Every worker with mainly one number')
    print('Version 3: Equal distribution of all numbers')
    version = input("Input: ")
    print('\nPlease enter a Number to select a Method:')
    print('Method 1: pertubation data')
    print('Method 2: corruption data')
    print('Method 3: normal')
    print('Method 4: opponent')
    method = input("Input: ")
    print('\nIs This a documentation Run? (y / press enter):')
    doc = (True if input("Input: ").lower() == 'y' else False)
    if doc: # Legt die Anzahl der Runs fest
        runs = int(input('Please enter a number of runs: '))
        if runs <1 or runs > 100:
            print("invalid number of runs")
            return

    print('Starting...') #So i know wether the programm screwed up or not

    #Documentation
    if(doc):
        wb = load_workbook(path+'Version'+str(version)+'.xlsx')
    else:
        runs = 1

    for runIndex in range(runs):
        start = time.time()

        print("\n\nRun Number "+str(runIndex))
        print(":::::::::::::::::::::::::::::::::::::::::::::::::::::::")
        print(":::::::::::::::::::::::::::::::::::::::::::::::::::::::")
        if(doc):
            column = get_column_letter(getColumn(wb['Average loss during Training']))  # Nehme einfach das erst beste Sheet und gucke was die aktuelle Spalte/ was der aktuelle run ist

        torch.manual_seed(np.round(time.time() * 1000))

        # check if GPU is available, otherwise use CPU
        #device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        device = torch.device("cpu")

        transform = transforms.Compose([transforms.ToTensor(), transforms.Normalize((0.1307,), (0.3081,))])

        # load MNIST dataset
        dataset_training = MNIST(root='./rootMNIST', train=True, download=True, transform=transform)
        #dataset_training2 = MNIST(root='./rootMNIST', train=True, download=True, transform=transform)

        dataset_testing = MNIST(root='./rootMNIST', train=False, download=True, transform=transform)

        # get the labels of each tuple (image, label) in the dataset
        labels = dataset_training.train_labels

        trainloader = DataLoader(dataset_training, batch_size=batchSize, shuffle=True)


        #Load Network and Autoencoder (to retrain delete global_model)
        try:
            global_model = torch.load((os.path.dirname(os.path.realpath(__file__)) + ("\\global_model")))
            global_model.eval()
            training = False
            print("Loading of  global_model successfull.")
        except FileNotFoundError:
            print("No saved global_model: Created new global model for Training.")
            global_model = NeuralNetwork().to(device)
            training = True

        try:
            autoencoder_disabled = False
            autoencoder = torch.load(os.path.dirname(os.path.realpath(__file__)) + "\\AutoEncoder")
            autoencoder.eval()
            print("Loading of Autoencoder successfull.")
        except FileNotFoundError and _pickle.UnpicklingError:
            if(training==False):
                autoencoder_disabled = False
                print("No saved global_model: Starting Training of Autoencoder.")
                autoencoder = prepareAutoEncoder(trainloader, device, global_model)
                print("Autoencoder trained.")
            else:
                autoencoder_disabled=True
                print("Autoencoder disabled, try again after training and saving global_model.")

        if version == '1':
            # Version 1
            # one worker has the majority of one number but still a few pictures of other numbers

            # get the indices of all tuples, where the label is '0'
            label_zero_indices = (labels == 0).nonzero()

            # from tensor to list with flatten().tolist()
            label_zero_indices = label_zero_indices.flatten().tolist()
            # print(len(label_zero_indices)) = 5923

            # get a subset of the dataset with the filtered indices
            label_zero_subset = torch.utils.data.Subset(dataset_training, label_zero_indices)

            # random_split expects a dataset and the wanted length, returns 10 non-overlapping new datasets, meaning a list of datasets?
            partition_of_training_data = torch.utils.data.random_split(label_zero_subset,
                                                                       [5023, 100, 100, 100, 100, 100, 100, 100, 100,
                                                                        100])  # 5023 + 9*100 = 5923

            # make a subset of the mnist dataset without zeroes: indices of all tuples where the label is NOT '0'
            label_not_zero_indices = (labels != 0).nonzero()
            label_not_zero_indices = label_not_zero_indices.flatten().tolist()
            # print(len(label_not_zero_indices)) = 54077

            # make the second subset without zeroes
            label_not_zero_subset = torch.utils.data.Subset(dataset_training, label_not_zero_indices)
            partition_of_training_data2 = torch.utils.data.random_split(label_not_zero_subset,
                                                                        [977, 5900, 5900, 5900, 5900, 5900, 5900, 5900,
                                                                         5900, 5900])  # 977 + 9*5900 = 54077

            # use ConcatDataset to join a list of Datasets
            new_dataset_list = []
            for i in range(10):
                new_dataset_list.append(
                    torch.utils.data.ConcatDataset([partition_of_training_data[i], partition_of_training_data2[i]]))

        elif version == '2':
            # get the indices of the numbers
            label_zero_indices = (labels == 0).nonzero()
            label_one_indices = (labels == 1).nonzero()
            label_two_indices = (labels == 2).nonzero()
            label_three_indices = (labels == 3).nonzero()
            label_four_indices = (labels == 4).nonzero()
            label_five_indices = (labels == 5).nonzero()
            label_six_indices = (labels == 6).nonzero()
            label_seven_indices = (labels == 7).nonzero()
            label_eight_indices = (labels == 8).nonzero()
            label_nine_indices = (labels == 9).nonzero()

            # from tensor to list
            label_zero_indices = label_zero_indices.flatten().tolist()
            label_one_indices = label_one_indices.flatten().tolist()
            label_two_indices = label_two_indices.flatten().tolist()
            label_three_indices = label_three_indices.flatten().tolist()
            label_four_indices = label_four_indices.flatten().tolist()
            label_five_indices = label_five_indices.flatten().tolist()
            label_six_indices = label_six_indices.flatten().tolist()
            label_seven_indices = label_seven_indices.flatten().tolist()
            label_eight_indices = label_eight_indices.flatten().tolist()
            label_nine_indices = label_nine_indices.flatten().tolist()

            # make subsets for each number
            label_zero_subset = torch.utils.data.Subset(dataset_training, label_zero_indices)
            label_one_subset = torch.utils.data.Subset(dataset_training, label_one_indices)
            label_two_subset = torch.utils.data.Subset(dataset_training, label_two_indices)
            label_three_subset = torch.utils.data.Subset(dataset_training, label_three_indices)
            label_four_subset = torch.utils.data.Subset(dataset_training, label_four_indices)
            label_five_subset = torch.utils.data.Subset(dataset_training, label_five_indices)
            label_six_subset = torch.utils.data.Subset(dataset_training, label_six_indices)
            label_seven_subset = torch.utils.data.Subset(dataset_training, label_seven_indices)
            label_eight_subset = torch.utils.data.Subset(dataset_training, label_eight_indices)
            label_nine_subset = torch.utils.data.Subset(dataset_training, label_nine_indices)

            # get the count for each number
            # print(len(label_zero_indices)) = 5923
            # print(len(label_one_indices)) = 6742
            # print(len(label_two_indices)) = 5958
            # print(len(label_three_indices)) = 6131
            # print(len(label_four_indices)) = 5842
            # print(len(label_five_indices)) = 5421
            # print(len(label_six_indices)) = 5918
            # print(len(label_seven_indices)) = 6265
            # print(len(label_eight_indices)) = 5851
            # print(len(label_nine_indices)) = 5949

            # split the numbers for the workers
            partition_of_training_data0 = torch.utils.data.random_split(label_zero_subset,
                                                                        [5023, 100, 100, 100, 100, 100, 100, 100, 100, 100])
            partition_of_training_data1 = torch.utils.data.random_split(label_one_subset,
                                                                        [100, 5842, 100, 100, 100, 100, 100, 100, 100, 100])
            partition_of_training_data2 = torch.utils.data.random_split(label_two_subset,
                                                                        [100, 100, 5058, 100, 100, 100, 100, 100, 100, 100])
            partition_of_training_data3 = torch.utils.data.random_split(label_three_subset,
                                                                        [100, 100, 100, 5231, 100, 100, 100, 100, 100, 100])
            partition_of_training_data4 = torch.utils.data.random_split(label_four_subset,
                                                                        [100, 100, 100, 100, 4942, 100, 100, 100, 100, 100])
            partition_of_training_data5 = torch.utils.data.random_split(label_five_subset,
                                                                        [100, 100, 100, 100, 100, 4521, 100, 100, 100, 100])
            partition_of_training_data6 = torch.utils.data.random_split(label_six_subset,
                                                                        [100, 100, 100, 100, 100, 100, 5018, 100, 100, 100])
            partition_of_training_data7 = torch.utils.data.random_split(label_seven_subset,
                                                                        [100, 100, 100, 100, 100, 100, 100, 5365, 100, 100])
            partition_of_training_data8 = torch.utils.data.random_split(label_eight_subset,
                                                                        [100, 100, 100, 100, 100, 100, 100, 100, 4951, 100])
            partition_of_training_data9 = torch.utils.data.random_split(label_nine_subset,
                                                                        [100, 100, 100, 100, 100, 100, 100, 100, 100, 5049])

            # concat datasets
            new_dataset_list = []
            for i in range(10):
                new_dataset_list.append(torch.utils.data.ConcatDataset(
                    [partition_of_training_data0[i], partition_of_training_data1[i], partition_of_training_data2[i],
                     partition_of_training_data3[i], partition_of_training_data4[i], partition_of_training_data5[i],
                     partition_of_training_data6[i], partition_of_training_data7[i], partition_of_training_data8[i],
                     partition_of_training_data9[i]]))

        elif version == '3':
            # Dividing the training data into num_clients, with each client having equal number of images
            new_dataset_list = torch.utils.data.random_split(dataset_training,
                                                             [int(dataset_training.data.shape[0] / numberOfClients) for _ in
                                                              range(numberOfClients)])

        elif version == '4':
            #Version 4
            #one worker has everything wrong and the others are untouched
            new_dataset_list = torch.utils.data.random_split(dataset_training,
                                                             [int(dataset_training.data.shape[0] / numberOfClients) for _ in
                                                              range(numberOfClients)])

        else:
            exit('Number should be 1, 2, 3 or 4')

        # for partX in partition gets 10x train_loader, one for each worker
        # train_loader is a list of DataLoaders, using new_dataset_list instead of partition_of_training_data loads the individually distributed datasets
        train_loader = [DataLoader(partX, batch_size=batchSize, shuffle=True) for partX in new_dataset_list]

        if version == '4':
            dataset_corrupt = Dataset_RandomOther(transform)
            train_loader[0] = DataLoader(dataset_corrupt, batch_size=batchSize, shuffle=True)

        # further divide data of clients into train and test
        training_loader__local_client = []
        testing_loader__local_client = []
        for i in range(numberOfClients):
            clientDataset = new_dataset_list[i]
            #20 percent of data given to client is used for testing and the rest for training
            training_data_c, testing_data_c = train_test_split(clientDataset, test_size=0.2,shuffle=True,random_state=42)
            training_loader__local_client.append(DataLoader(training_data_c, batch_size=batchSize, shuffle=True))
            testing_loader__local_client.append(DataLoader(testing_data_c, batch_size=batchSize, shuffle=False))

        test_loader = DataLoader(dataset_testing, batch_size=batchSize, shuffle=False)

        print(" Number of Elements in Training Dataset:", len(dataset_training))
        print(" Number of Elements in Test Dataset:", len(dataset_testing))
        # initialize global server model and identical client models

        client_models = [NeuralNetwork().to(device) for _ in range(numberOfClients)]
        print(" Model:", global_model)

        for model in client_models:
            model.load_state_dict(global_model.state_dict())
        # initialisation of Stochastic Gradient Descent as optimizer and  StepLR as scheduler for all client models
        optimizer = [torch.optim.AdamW(model.parameters(), learningrate) for model in client_models]
        scheduler = [StepLR(optim, step_size=1) for optim in optimizer]

        # waiting 0.1s so that prints can finish before tqdm
        time.sleep(0.1)

        # Starting Training on regular trainmethod, then proceed to train on salt-and-pepper and gaussian noise to make network more robust
        # TODO:Add Train Methods for Gaussian Noise and Speckle Noise to TrainingMethods-List
        TrainingMethods = [train_client,train_client_SaltPepper,train_client_HardSaltPepper]
        i=1
        for currentTrainingMethod in TrainingMethods:
            if(i>1):
                numberOfRounds=5
            else:
                numberOfRounds=20
            if training:
                print("")
                print("Training Network on Method ", currentTrainingMethod)
                print("")
                time.sleep(0.1)
            for round in range(numberOfRounds):
                # to speed up training only train 4 randomly selected clients
                selectedClients = np.random.permutation(numberOfClients)[:numberOfSelectedClients]


                # train selected clients
                current_round_loss = 0
                if (training):
                    for i in tqdm(range(numberOfSelectedClients), position=0, leave=True):
                        current_round_loss += currentTrainingMethod(client_models[selectedClients[i]], device, optimizer[selectedClients[i]], training_loader__local_client[selectedClients[i]],
                                                           epoch=numberOfEpochs)

                    print("Clients that have been trained:",selectedClients[0],",",selectedClients[1],", ",selectedClients[2],", ",selectedClients[3])

                    # for each client test its model with test_loader of every client to get a 10x10 matrix
                    tensorList = []
                    clientid2 = 0 # documentation
                    for clientid in range(numberOfClients):
                        results = []
                        for testi in (testing_loader__local_client):
                            test_loss, acc = test(client_models[clientid], device, testi)
                            results.append(acc)
                            # Documentation
                            if(doc):
                                wb['CM ' + str(clientid2) + str(clientid)][column + str(round + 2)].value = acc
                                clientid2 = (clientid2 + 1) % numberOfClients
                        tensorList.append(results)
                    torch.set_printoptions(linewidth=200)
                    tensorList = torch.tensor(tensorList)
                    print("Local Accuracy Matrix :")
                    print(tensorList)
                    # Documentation
                    #if (doc):
                    #    for clientid in range(numberOfClients):
                    #        for testi in (testing_loader__local_client):
                    #            print('CM '+str(testi)+str(clientid))
                    #            print(column + str(round + 2))
                    #            print(clientid)
                    #            print(testi)
                    #            wb['CM '+str(testi)+str(clientid)][column + str(round + 2)].value = tensorList[clientid][testi]

                    # aggregate results of client training and update global- and all client models
                    server_aggregate(global_model, client_models,selectedClients)



                # test current state of updated global model
                test_loss, accuracy = test(global_model, device, test_loader)
                # Documentation
                if(doc):
                    wb['Average loss during Training'][column+str(round+2)].value = current_round_loss / numberOfSelectedClients
                    wb['Global Test loss'][column + str(round + 2)].value = test_loss
                    wb['Global Accuracy'][column + str(round + 2)].value = accuracy
                if (training):
                    print('Round %d :' % (round + 1),
                          'Original Image: Average loss during Training: %0.3g | Global Test loss: %0.3g | Global Accuracy: %0.3f' % (
                          current_round_loss / numberOfSelectedClients, test_loss, accuracy))
                else:
                    print('Round %d :' % (round + 1),
                          'Original Image:  Global Test loss: %0.3g | Global Accuracy: %0.3f' % (test_loss, accuracy))
                test_loss, accuracy = test_SaltPepper(global_model, device, test_loader)
                if (training):
                    print('Round %d :' % (round + 1),
                          'Image with salt-and-pepper-noise without Denoising Autoencoder : Global Test loss: %0.3g | Global Accuracy: %0.3f' % (
                               test_loss, accuracy))
                else:
                    print('Round %d :' % (round + 1),
                          'Image with salt-and-pepper-noise without Denoising Autoencoder :  Global Test loss: %0.3g | Global Accuracy: %0.3f' % (
                          test_loss, accuracy))

                test_loss, accuracy = test_HardSaltPepper(global_model, device, test_loader)
                if (training):
                    print('Round %d :' % (round + 1),
                          'Image with hard salt-and-pepper-noise without Denoising Autoencoder : Global Test loss: %0.3g | Global Accuracy: %0.3f' % (
                              test_loss, accuracy))
                else:
                    print('Round %d :' % (round + 1),
                          'Image with hard salt-and-pepper-noise without Denoising Autoencoder :  Global Test loss: %0.3g | Global Accuracy: %0.3f' % (
                              test_loss, accuracy))

                test_loss, accuracy = test_SaltPepperWithAutoEncoder(global_model, device, test_loader, autoencoder)
                if (training==True and autoencoder_disabled==False):
                    print('Round %d :' % (round + 1),
                          'Image with salt-and-pepper-noise with pre-processing by Denoising Autoencoder :  Global Test loss: %0.3g | Global Accuracy: %0.3f' % (
                               test_loss, accuracy))
                elif(autoencoder_disabled==False):
                    print('Round %d :' % (round + 1),
                          'Image with salt-and-pepper-noise with pre-processing by Denoising Autoencoder :  Global Test loss: %0.3g | Global Accuracy: %0.3f' % (
                              test_loss, accuracy))
                test_loss, accuracy = test_HardSaltPepperWithAutoencoder(global_model, device, test_loader, autoencoder)
                if (training == True and autoencoder_disabled == False):
                    print('Round %d :' % (round + 1),
                          'Image with hard salt-and-pepper-noise with pre-processing by Denoising Autoencoder :  Global Test loss: %0.3g | Global Accuracy: %0.3f' % (
                              test_loss, accuracy))
                elif (autoencoder_disabled == False):
                    print('Round %d :' % (round + 1),
                          'Image with hard salt-and-pepper-noise with pre-processing by Denoising Autoencoder :  Global Test loss: %0.3g | Global Accuracy: %0.3f' % (
                              test_loss, accuracy))

            #TODO: Include noise accuracies into table?
                i+=1
                if (training):
                    # adapt learning rates with scheduler
                    for i in range(len(scheduler)):
                        if i in selectedClients:
                            scheduler[i].step()

                # wait 0.1s so that prints can finish
                time.sleep(0.1)

                #Documentation
                if(doc):
                    wb.save(path+'Version'+str(version)+'.xlsx')

            # Time orientation
            print("________________________")
            print("Time spent on this run: "+str((time.time()-start)/60)+"min")

            if(training):
                #Save classification model after training
                torch.save(global_model, (os.path.dirname(os.path.realpath(__file__)) + ("\\global_model")))


if __name__ == '__main__':
    main()